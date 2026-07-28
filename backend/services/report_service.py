from datetime import datetime
from services.market_service import MarketService
from services.ai_service import AIService
from services.news_service import NewsService

class ReportService:
    def __init__(self):
        self._market = MarketService()
        self._ai     = AIService()
        self._news   = NewsService()

    def get_report_data(self):
        return {
            "generated_at": datetime.utcnow().isoformat(),
            "prices":       self._market.get_current_prices(),
            "summary":      self._market.get_summary(),
            "forecast":     self._ai.quick_forecast(),
            "sentiment":    self._news.get_aggregate_sentiment(),
            "history_30d":  self._market.get_history_with_ma(30),
        }

    # ── PDF ───────────────────────────────────────────────────────────────────
    def generate_pdf(self) -> bytes:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.units import cm
            import io

            data   = self.get_report_data()
            prices = data["prices"]
            summ   = data["summary"]
            fc     = data["forecast"]
            sent   = data["sentiment"]
            buf    = io.BytesIO()

            doc    = SimpleDocTemplate(buf, pagesize=A4,
                                       topMargin=2*cm, bottomMargin=2*cm,
                                       leftMargin=2*cm, rightMargin=2*cm)
            styles = getSampleStyleSheet()
            story  = []

            # Título
            title_style = ParagraphStyle("title", parent=styles["Title"],
                                         fontSize=22, textColor=colors.HexColor("#1a1a1a"))
            story.append(Paragraph("lauOIL — Relatório de Mercado", title_style))
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph(f"Gerado em: {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}",
                                   styles["Normal"]))
            story.append(Spacer(1, 0.8*cm))

            # Secção: Preços
            story.append(Paragraph("1. Preços Actuais", styles["Heading1"]))
            price_data = [
                ["Commodity", "Preço (USD)", "Moeda", "Unidade"],
                ["Brent",       f"${prices['brent']['price']}",       "USD", "barril"],
                ["WTI",         f"${prices['wti']['price']}",         "USD", "barril"],
                ["Gás Natural", f"${prices['natural_gas']['price']}","USD", "mmbtu"],
                ["Spread B-W",  f"${prices['spread']}",              "USD", "barril"],
            ]
            tbl = Table(price_data, colWidths=[4*cm, 4*cm, 3*cm, 4*cm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#f59e0b")),
                ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f9f9f9")]),
                ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#e5e5e5")),
                ("FONTSIZE",    (0,0), (-1,-1), 10),
                ("PADDING",     (0,0), (-1,-1), 8),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 0.8*cm))

            # Secção: Variações
            story.append(Paragraph("2. Variações de Mercado", styles["Heading1"]))
            chg = summ["changes"]
            var_data = [
                ["Período",  "Variação (%)"],
                ["Diária",   f"{chg['daily']:+}%"],
                ["Semanal",  f"{chg['weekly']:+}%"],
                ["Mensal",   f"{chg['monthly']:+}%"],
                ["Anual",    f"{chg['yearly']:+}%"],
            ]
            tbl2 = Table(var_data, colWidths=[6*cm, 6*cm])
            tbl2.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#1a1a1a")),
                ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f9f9f9")]),
                ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#e5e5e5")),
                ("FONTSIZE",    (0,0), (-1,-1), 10),
                ("PADDING",     (0,0), (-1,-1), 8),
            ]))
            story.append(tbl2)
            story.append(Spacer(1, 0.8*cm))

            # Secção: Previsão IA
            story.append(Paragraph("3. Previsão por Inteligência Artificial", styles["Heading1"]))
            story.append(Paragraph(
                f"Preço previsto a 7 dias: <b>${fc['price_7d']}</b> | "
                f"30 dias: <b>${fc['price_30d']}</b> | "
                f"Confiança: <b>{fc['confidence']}%</b> | "
                f"Sinal: <b>{fc['signal'].replace('_',' ').upper()}</b>",
                styles["Normal"]
            ))
            story.append(Spacer(1, 0.8*cm))

            # Secção: Sentimento
            story.append(Paragraph("4. Análise de Sentimento do Mercado", styles["Heading1"]))
            story.append(Paragraph(
                f"Score: <b>{sent['score']:+}</b> ({sent['label'].upper()}) | "
                f"Positivo: {sent['positive_pct']}% | "
                f"Negativo: {sent['negative_pct']}% | "
                f"Artigos analisados: {sent['articles_analyzed']}",
                styles["Normal"]
            ))
            story.append(Spacer(1, 1*cm))

            # Rodapé
            story.append(Paragraph(
                "lauOIL v1.0 — Plataforma de Inteligência Artificial para o Mercado Petrolífero",
                ParagraphStyle("footer", parent=styles["Normal"],
                               fontSize=8, textColor=colors.grey)
            ))

            doc.build(story)
            return buf.getvalue()

        except ImportError:
            return b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n3 0 obj<</Type/Page/MediaBox[0 0 595 842]/Parent 2 0 R>>endobj\nxref\n0 4\n0000000000 65535 f\n0000000009 00000 n\n0000000058 00000 n\n0000000115 00000 n\ntrailer<</Size 4/Root 1 0 R>>\nstartxref\n190\n%%EOF"

    # ── Excel ─────────────────────────────────────────────────────────────────
    def generate_excel(self) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            import io

            data    = self.get_report_data()
            prices  = data["prices"]
            summ    = data["summary"]
            history = data["history_30d"]["data"]
            fc      = data["forecast"]

            wb = openpyxl.Workbook()

            # ── Folha 1: Resumo ──
            ws1 = wb.active
            ws1.title = "Resumo"
            amber  = PatternFill("solid", fgColor="F59E0B")
            dark   = PatternFill("solid", fgColor="1A1A1A")
            hdr_ft = Font(bold=True, color="FFFFFF")
            thin   = Border(
                left=Side(style="thin", color="E5E5E5"),
                right=Side(style="thin", color="E5E5E5"),
                top=Side(style="thin", color="E5E5E5"),
                bottom=Side(style="thin", color="E5E5E5"),
            )

            ws1["A1"] = "lauOIL — Resumo de Mercado"
            ws1["A1"].font = Font(bold=True, size=16)
            ws1["A2"] = f"Gerado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}"

            ws1["A4"] = "Commodity";  ws1["B4"] = "Preço (USD)"
            for cell in ws1["4:4"]:
                cell.fill = amber; cell.font = hdr_ft

            ws1["A5"]="Brent";        ws1["B5"]=prices["brent"]["price"]
            ws1["A6"]="WTI";          ws1["B6"]=prices["wti"]["price"]
            ws1["A7"]="Gás Natural";  ws1["B7"]=prices["natural_gas"]["price"]
            ws1["A8"]="Spread B-W";   ws1["B8"]=prices["spread"]

            ws1["A10"]="Período";     ws1["B10"]="Variação (%)"
            for cell in ws1["10:10"]:
                cell.fill = dark; cell.font = hdr_ft

            ws1["A11"]="Diária";  ws1["B11"]=summ["changes"]["daily"]
            ws1["A12"]="Semanal"; ws1["B12"]=summ["changes"]["weekly"]
            ws1["A13"]="Mensal";  ws1["B13"]=summ["changes"]["monthly"]
            ws1["A14"]="Anual";   ws1["B14"]=summ["changes"]["yearly"]

            ws1["A16"]="Previsão IA (7d)";  ws1["B16"]=fc["price_7d"]
            ws1["A17"]="Previsão IA (30d)"; ws1["B17"]=fc["price_30d"]
            ws1["A18"]="Confiança (%)";     ws1["B18"]=fc["confidence"]
            ws1["A19"]="Sinal";             ws1["B19"]=fc["signal"].replace("_"," ").upper()

            ws1.column_dimensions["A"].width = 22
            ws1.column_dimensions["B"].width = 18

            # ── Folha 2: Histórico ──
            ws2 = wb.create_sheet("Histórico 30d")
            headers = ["Data", "Brent", "WTI", "MA7", "MA20", "MA50", "Volume"]
            for col, h in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.fill = amber; cell.font = hdr_ft

            for row_i, rec in enumerate(history, 2):
                ws2.cell(row=row_i, column=1, value=rec["date"])
                ws2.cell(row=row_i, column=2, value=rec["brent"])
                ws2.cell(row=row_i, column=3, value=rec["wti"])
                ws2.cell(row=row_i, column=4, value=rec.get("ma7"))
                ws2.cell(row=row_i, column=5, value=rec.get("ma20"))
                ws2.cell(row=row_i, column=6, value=rec.get("ma50"))
                ws2.cell(row=row_i, column=7, value=rec.get("volume"))

            for col in range(1, 8):
                ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        except ImportError:
            return b""
